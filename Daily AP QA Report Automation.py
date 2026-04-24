from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta
import time
import os
import json
import base64
import tkinter as tk
from tkinter import ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule


class UnifiedAutomation:
    def __init__(self, download_path=None):
        """Unified initialization for both automations"""
        self.cache_file = os.path.join(os.path.expanduser("~"), '.orderbahn_cache.json')
        
        if download_path is None:
            download_path = os.path.join(os.path.expanduser("~"), 'downloads')
        os.makedirs(download_path, exist_ok=True)
        
        self.download_path = download_path
        self.driver = None
        self.selected_tenant = None
        self.credentials = {}
        
        # Downloaded file paths
        self.orderbahn_csv = None
        self.erp_csv = None
        
        # Progress tracking
        self.current_step = 0
        self.total_steps = 8
        self.completed_steps = []
        self.failed_steps = []
        
        self.all_tenants = [
            "Ci Select",
            "Core Office Interiors",
            "Creative Office Resources",
            "Intivity",
            "Leland Furniture",
            "Office Creations",
            "Office Resources Inc.",
            "Office Revolution",
            "Ofi",
            "Op Houston",
            "Storey Kenworthy",
            "Wurkwel"
        ]
        
        self.erp_urls = {
            "Office Revolution": "https://officerevolution.coreincloud.com/core/login ",
            "Creative Office Resources": "https://creativeofficeresources.coreincloud.com/core/login ",
            "Core Office Interiors": "https://coreofficeinteriors.coreincloud.com/core/login ",
            "Op Houston": "https://ophoustonwrg.coreincloud.com/core/login ",
            "Office Creations": "https://officecreations.coreincloud.com/core/login ",
        }
    
    def init_driver(self, browser="chrome"):
        """Initialize web driver with browser compatibility"""
        if browser.lower() == "chrome":
            options = Options()
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-notifications')
            options.page_load_strategy = 'eager'
            
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            prefs = {
                "download.default_directory": self.download_path,
                "download.prompt_for_download": False,
                "profile.default_content_setting_values.notifications": 2
            }
            options.add_experimental_option("prefs", prefs)
            
            print(f"Starting {browser.title()}...")
            self.driver = webdriver.Chrome(options=options)
            
        elif browser.lower() == "firefox":
            from selenium.webdriver.firefox.options import Options as FirefoxOptions
            options = FirefoxOptions()
            options.set_preference("browser.download.folderList", 2)
            options.set_preference("browser.download.dir", self.download_path)
            options.set_preference("browser.download.useDownloadDir", True)
            options.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv")
            
            print(f"Starting {browser.title()}...")
            self.driver = webdriver.Firefox(options=options)
            
        elif browser.lower() == "edge":
            from selenium.webdriver.edge.options import Options as EdgeOptions
            options = EdgeOptions()
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            
            prefs = {
                "download.default_directory": self.download_path,
                "download.prompt_for_download": False
            }
            options.add_experimental_option("prefs", prefs)
            
            print(f"Starting {browser.title()}...")
            self.driver = webdriver.Edge(options=options)
        
        self.driver.set_page_load_timeout(90)
        self.driver.set_script_timeout(30)
        print(f"{browser.title()} started successfully\n")
    
    def close_driver(self):
        """Close the driver"""
        if self.driver:
            self.driver.quit()
            self.driver = None
    
    def load_cache(self):
        """Load cached credentials and preferences"""
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r') as f:
                    cache = json.load(f)
                    if 'password' in cache:
                        cache['password'] = base64.b64decode(cache['password'].encode()).decode()
                    if 'erp_password' in cache:
                        cache['erp_password'] = base64.b64decode(cache['erp_password'].encode()).decode()
                    print("[OK] Cache loaded\n")
                    return cache
        except Exception as e:
            print(f"Could not load cache: {e}")
        return {}
    
    def save_cache(self, email, password, erp_username=None, erp_password=None, tenant_name=None, skip_mfa=True):
        """Save all credentials and preferences to cache"""
        try:
            cache = {
                'email': email,
                'password': base64.b64encode(password.encode()).decode(),
                'tenant_name': tenant_name,
                'skip_mfa': skip_mfa,
                'last_used': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            if erp_username:
                cache['erp_username'] = erp_username
            if erp_password:
                cache['erp_password'] = base64.b64encode(erp_password.encode()).decode()
            
            with open(self.cache_file, 'w') as f:
                json.dump(cache, f)
            print("[OK] Credentials cached for next time\n")
        except Exception as e:
            print(f"⚠ Could not save cache: {e}\n")
    
    def show_status_window(self, title, message, step_num=0, duration=3000):
        """Show non-blocking status window that auto-closes"""
        root = tk.Tk()
        root.withdraw()
        
        status_win = tk.Toplevel(root)
        status_win.title(title)
        status_win.geometry("500x200")
        status_win.attributes('-topmost', True)
        
        # Remove window decorations and make it non-blocking
        status_win.overrideredirect(True)
        
        # Center the window
        status_win.update_idletasks()
        width = status_win.winfo_width()
        height = status_win.winfo_height()
        x = (status_win.winfo_screenwidth() // 2) - (width // 2)
        y = (status_win.winfo_screenheight() // 2) - (height // 2)
        status_win.geometry(f'{width}x{height}+{x}+{y}')
        
        # Add a frame for styling
        frame = ttk.Frame(status_win, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = tk.Label(frame, text=title, font=("Arial", 14, "bold"), fg="#2196F3")
        title_label.pack(pady=(10, 15))
        
        # Message
        msg_label = tk.Label(frame, text=message, font=("Arial", 11), wraplength=400, justify=tk.LEFT)
        msg_label.pack(pady=10)
        
        # Progress indicator
        if step_num > 0:
            self.current_step = step_num
            progress_text = f"Step {step_num} of {self.total_steps}"
            progress_label = tk.Label(frame, text=progress_text, font=("Arial", 10, "italic"), fg="#666")
            progress_label.pack(pady=5)
            
            remaining = self.total_steps - step_num
            remaining_text = f"Remaining steps: {remaining}"
            remaining_label = tk.Label(frame, text=remaining_text, font=("Arial", 9), fg="#888")
            remaining_label.pack(pady=5)
        
        # Auto-close after duration
        status_win.after(duration, lambda: status_win.destroy())
        
        # Force update to show window immediately
        status_win.update()
        root.update()
        
        # Store reference to close it early if needed
        self.status_window = status_win
        self.status_root = root
    
    def close_status_window(self):
        """Close the status window if it's open"""
        if hasattr(self, 'status_window') and self.status_window:
            try:
                self.status_window.destroy()
                self.status_root.destroy()
            except:
                pass
            self.status_window = None
            self.status_root = None
    
    def show_initial_setup_dialog(self, cache):
        """Show comprehensive setup dialog"""
        root = tk.Tk()
        root.title("OrderBahn & ERP Automation Setup")
        root.geometry("600x750")
        root.attributes('-topmost', True)
        
        results = {
            'cancelled': True,
            'orderbahn_email': '',
            'orderbahn_password': '',
            'tenant': '',
            'erp_username': '',
            'erp_password': '',
            'save_credentials': True,
            'browser': 'chrome'
        }
        
        main_frame = ttk.Frame(root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        title_label = tk.Label(main_frame, text="Automation Setup", 
                              font=("Arial", 16, "bold"), fg="#2196F3")
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        row = 1
        
        section1 = tk.Label(main_frame, text="OrderBahn Credentials", 
                           font=("Arial", 12, "bold"), fg="#333")
        section1.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(10, 5))
        row += 1
        
        tk.Label(main_frame, text="Email:", font=("Arial", 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        email_entry = ttk.Entry(main_frame, width=40)
        email_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        if cache.get('email'):
            email_entry.insert(0, cache['email'])
        row += 1
        
        tk.Label(main_frame, text="Password:", font=("Arial", 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        password_entry = ttk.Entry(main_frame, width=40, show='*')
        password_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        if cache.get('password'):
            password_entry.insert(0, cache['password'])
        row += 1
        
        section2 = tk.Label(main_frame, text="Tenant Selection", 
                           font=("Arial", 12, "bold"), fg="#333")
        section2.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(20, 5))
        row += 1
        
        tk.Label(main_frame, text="Select Tenant:", font=("Arial", 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        tenant_var = tk.StringVar()
        tenant_combo = ttk.Combobox(main_frame, textvariable=tenant_var, width=38, state='readonly')
        tenant_combo['values'] = self.all_tenants
        tenant_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        
        if cache.get('tenant_name') and cache['tenant_name'] in self.all_tenants:
            tenant_combo.set(cache['tenant_name'])
        elif self.all_tenants:
            tenant_combo.set(self.all_tenants[0])
        row += 1
        
        section3 = tk.Label(main_frame, text="ERP Credentials (if applicable)", 
                           font=("Arial", 12, "bold"), fg="#333")
        section3.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(20, 5))
        row += 1
        
        erp_info = tk.Label(main_frame, 
                           text="Only needed if your tenant has an ERP system\n(Leave blank if not applicable)", 
                           font=("Arial", 9), fg="gray")
        erp_info.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=5)
        row += 1
        
        tk.Label(main_frame, text="ERP Username:", font=("Arial", 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        erp_user_entry = ttk.Entry(main_frame, width=40)
        erp_user_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        if cache.get('erp_username'):
            erp_user_entry.insert(0, cache['erp_username'])
        row += 1
        
        tk.Label(main_frame, text="ERP Password:", font=("Arial", 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        erp_pass_entry = ttk.Entry(main_frame, width=40, show='*')
        erp_pass_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        if cache.get('erp_password'):
            erp_pass_entry.insert(0, cache['erp_password'])
        row += 1
        
        section4 = tk.Label(main_frame, text="Browser Selection", 
                           font=("Arial", 12, "bold"), fg="#333")
        section4.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(20, 5))
        row += 1
        
        tk.Label(main_frame, text="Browser:", font=("Arial", 10)).grid(row=row, column=0, sticky=tk.W, pady=5)
        browser_var = tk.StringVar(value="chrome")
        browser_frame = ttk.Frame(main_frame)
        browser_frame.grid(row=row, column=1, sticky=tk.W, pady=5)
        ttk.Radiobutton(browser_frame, text="Chrome", variable=browser_var, value="chrome").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(browser_frame, text="Firefox", variable=browser_var, value="firefox").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(browser_frame, text="Edge", variable=browser_var, value="edge").pack(side=tk.LEFT, padx=5)
        row += 1
        
        save_var = tk.BooleanVar(value=True)
        save_check = ttk.Checkbutton(main_frame, text="Save credentials for next time", 
                                     variable=save_var)
        save_check.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(20, 10))
        row += 1
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        def on_start():
            if not email_entry.get():
                tk.messagebox.showerror("Error", "OrderBahn email is required!")
                return
            if not password_entry.get():
                tk.messagebox.showerror("Error", "OrderBahn password is required!")
                return
            if not tenant_var.get():
                tk.messagebox.showerror("Error", "Please select a tenant!")
                return
            
            selected_tenant = tenant_var.get()
            if selected_tenant in self.erp_urls:
                if not erp_user_entry.get() or not erp_pass_entry.get():
                    response = tk.messagebox.askyesno(
                        "ERP Credentials Missing",
                        f"{selected_tenant} has an ERP system.\n\n"
                        "ERP credentials are empty. Do you want to continue without ERP automation?",
                        icon='warning'
                    )
                    if not response:
                        return
            
            results['cancelled'] = False
            results['orderbahn_email'] = email_entry.get()
            results['orderbahn_password'] = password_entry.get()
            results['tenant'] = tenant_var.get()
            results['erp_username'] = erp_user_entry.get()
            results['erp_password'] = erp_pass_entry.get()
            results['save_credentials'] = save_var.get()
            results['browser'] = browser_var.get()
            
            root.quit()
            root.destroy()
        
        def on_cancel():
            root.quit()
            root.destroy()
        
        start_btn = tk.Button(button_frame, text="Start Automation", command=on_start,
                             font=("Arial", 11, "bold"), bg="#4CAF50", fg="white",
                             padx=30, pady=8, cursor="hand2")
        start_btn.pack(side=tk.LEFT, padx=10)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel,
                              font=("Arial", 11), bg="#f44336", fg="white",
                              padx=30, pady=8, cursor="hand2")
        cancel_btn.pack(side=tk.LEFT, padx=10)
        
        root.update_idletasks()
        width = root.winfo_width()
        height = root.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        root.geometry(f'{width}x{height}+{x}+{y}')
        
        root.mainloop()
        return results
    
    def show_final_summary(self):
        """Show final summary of what was completed and what was skipped/failed"""
        # Close any open status window first
        self.close_status_window()
        
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        summary = "📊 AUTOMATION SUMMARY\n\n"
        
        if self.completed_steps:
            summary += "✅ COMPLETED:\n"
            for step in self.completed_steps:
                summary += f"   {step}\n"
            summary += "\n"
        
        if self.failed_steps:
            summary += "❌ FAILED:\n"
            for step in self.failed_steps:
                summary += f"   {step}\n"
            summary += "\n"
        
        summary += f"📁 All files saved to:\n{self.download_path}"
        
        tk.messagebox.showinfo("🎉 Automation Complete!", summary)
        root.destroy()
    
    def wait_for_page_load(self, timeout=15):
        """Wait for page to fully load"""
        for i in range(timeout):
            try:
                content_elements = self.driver.find_elements(By.XPATH, "//td | //div[contains(@class, 'content')]")
                loading_elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'Loading') or contains(@class, 'loading')]")
                
                if len(content_elements) > 5 and len(loading_elements) == 0:
                    return True
                time.sleep(1)
            except:
                time.sleep(1)
        return False
    
    def wait_for_download(self, timeout=30, file_prefix=None):
        """Wait for a new CSV file to appear in download folder"""
        print("Waiting for download to complete...")
        before = set(os.listdir(self.download_path))
        
        end_time = time.time() + timeout
        while time.time() < end_time:
            try:
                after = set(os.listdir(self.download_path))
                new_files = after - before
                csv_files = [f for f in new_files if f.endswith('.csv') and not f.endswith('.crdownload')]
                
                if csv_files:
                    csv_file = csv_files[0]
                    file_path = os.path.join(self.download_path, csv_file)
                    
                    initial_size = os.path.getsize(file_path)
                    time.sleep(1)
                    final_size = os.path.getsize(file_path)
                    
                    if initial_size == final_size and final_size > 0:
                        return csv_file
                
                time.sleep(1)
            except:
                time.sleep(1)
        
        return None
    
    def run_orderbahn_automation(self, email, password, tenant_name, skip_mfa=True):
        """Run OrderBahn automation with progress tracking"""
        try:
            print("=" * 60)
            print("PART 1: ORDERBAHN AUTOMATION")
            print("=" * 60 + "\n")
            
            # Step 1
            self.show_status_window("⏳ Step 1/8: Logging in...", "Please wait while we log in to OrderBahn.\n\nThis may take up to 30 seconds.", 1)
            print("[1/7] Logging in to OrderBahn...")
            self.driver.get("https://orderbahn.com/login ")
            time.sleep(3)
            
            email_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email']"))
            )
            email_field.send_keys(email)
            
            password_field = self.driver.find_element(By.XPATH, "//input[@type='password']")
            password_field.send_keys(password)
            
            submit_btn = self.driver.find_element(By.XPATH, "//button[@type='submit']")
            submit_btn.click()
            print("[OK] Login submitted\n")
            
            # Step 2
            self.show_status_window("⏳ Step 2/8: Handling Authentication...", "Waiting for MFA (if required) and tenant selection...", 2)
            print("[2/7] Handling authentication...")
            mfa_handled = False
            tenant_selected = False
            
            for i in range(30):
                time.sleep(1.5)
                try:
                    current_url = self.driver.current_url
                    
                    if ("setupmfa" in current_url or "mfa" in current_url.lower()) and not mfa_handled:
                        print("🔐 MFA page detected")
                        
                        if skip_mfa:
                            try:
                                skip_btn = WebDriverWait(self.driver, 3).until(
                                    EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'Skip MFA') or contains(text(),'Skip') or contains(text(),'Later')]"))
                                )
                                skip_btn.click()
                                print("✓ MFA skipped\n")
                                mfa_handled = True
                                time.sleep(2)
                                continue
                            except:
                                pass
                        
                        print("⚠ Cannot skip MFA - code required")
                        self.close_status_window()
                        root = tk.Tk()
                        root.withdraw()
                        root.attributes('-topmost', True)
                        mfa_code = tk.simpledialog.askstring("Enter MFA Code", 
                                                             "Enter the 6-digit MFA code:",
                                                             parent=root)
                        root.destroy()
                        
                        if mfa_code:
                            try:
                                mfa_input = WebDriverWait(self.driver, 5).until(
                                    EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @type='tel' or contains(@name, 'code')]"))
                                )
                                mfa_input.clear()
                                mfa_input.send_keys(mfa_code)
                                
                                verify_btn = self.driver.find_element(By.XPATH, "//button[contains(., 'Verify') or contains(., 'Submit') or contains(., 'Confirm')]")
                                verify_btn.click()
                                print("✓ MFA verification submitted\n")
                                mfa_handled = True
                                skip_mfa = False
                                time.sleep(3)
                            except Exception as e:
                                print(f"⚠ MFA submission error: {e}")
                        
                        mfa_handled = True
                    
                    elif "select-tenant" in current_url and not tenant_selected:
                        print(f"🏢 Auto-selecting tenant: {tenant_name}")
                        time.sleep(2)
                        
                        tenant_elements = self.driver.find_elements(By.XPATH, 
                            "//button[contains(@class, 'tenant-selection-button')]")
                        
                        if tenant_elements:
                            found_tenant = False
                            for elem in tenant_elements:
                                try:
                                    tenant_text = elem.find_element(By.XPATH, ".//h6").text.strip()
                                    
                                    if tenant_text == tenant_name:
                                        print(f"✓ Found and clicking: {tenant_name}")
                                        elem.click()
                                        time.sleep(1)
                                        
                                        try:
                                            login_btn = WebDriverWait(self.driver, 3).until(
                                                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Login')]"))
                                            )
                                            login_btn.click()
                                            print("✓ Login button clicked\n")
                                        except:
                                            pass
                                        
                                        found_tenant = True
                                        tenant_selected = True
                                        time.sleep(2)
                                        break
                                except:
                                    continue
                            
                            if not found_tenant:
                                print(f"❌ Could not find tenant: {tenant_name}")
                                raise Exception(f"Tenant '{tenant_name}' not found on page")
                        else:
                            print("❌ No tenant elements found")
                            raise Exception("No tenants available")
                    
                    elif "dashboard" in current_url or "bill" in current_url:
                        print("✓ Reached dashboard\n")
                        break
                        
                except Exception as e:
                    continue
            
            # Step 3
            self.show_status_window("⏳ Step 3/8: Navigating to Bills...", "Navigating to Bills page...", 3)
            print("[3/7] Navigating to Bills...")
            if "bill" not in self.driver.current_url:
                po_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//*[text()='PO Transactions']"))
                )
                po_btn.click()
                time.sleep(1)
                
                bill_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'Bill')]"))
                )
                bill_btn.click()
                print("✓ Navigated to Bills\n")
            
            # Step 4
            self.show_status_window("⏳ Step 4/8: Loading Bills Page...", "Loading Bills page data...", 4)
            print("[4/7] Loading Bills page...")
            self.wait_for_page_load(30)
            print("✓ Bills page loaded\n")
            
            # Step 5
            self.show_status_window("⏳ Step 5/8: Setting Filters...", "Setting up 'Daily qa' filter...", 5)
            print("[5/7] Setting up filters...")
            try:
                dropdown_selectors = [
                    "//div[@class='MuiSelect-select MuiSelect-outlined MuiInputBase-input MuiOutlinedInput-input']",
                    "//div[@role='button' and @aria-haspopup='listbox']",
                    "//div[contains(@class, 'MuiSelect-select')]"
                ]
                
                dropdown_clicked = False
                for selector in dropdown_selectors:
                    try:
                        dropdown = WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        dropdown.click()
                        print("✓ Dropdown clicked")
                        dropdown_clicked = True
                        time.sleep(4)
                        break
                    except:
                        continue
                
                if dropdown_clicked:
                    option_selectors = [
                        "//li[@role='option' and contains(., 'Daily qa')]",
                        "//li[contains(text(), 'Daily qa')]",
                        "//*[text()='Daily qa']"
                    ]
                    
                    for selector in option_selectors:
                        try:
                            option = WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, selector))
                            )
                            option.click()
                            print("✓ Daily qa selected")
                            time.sleep(15)
                            self.wait_for_page_load(30)
                            break
                        except:
                            continue
                print()
            except Exception as e:
                print(f"⚠ Filter error: {e}\n")
            
            # Step 6
            self.show_status_window("⏳ Step 6/8: Setting Date Filter...", "Setting date filter to yesterday...", 6)
            print("[6/7] Setting date filter...")
            yesterday = datetime.now() - timedelta(days=1)
            yesterday_str = yesterday.strftime("%m/%d/%Y 11:59 pm")
            print(f"Date: {yesterday_str}")
            
            try:
                filters_button_selectors = [
                    "//button[contains(., 'Filters') and contains(@class, 'MuiButton')]",
                    "//button[.//*[@data-testid='FilterListIcon']]"
                ]
                
                for selector in filters_button_selectors:
                    try:
                        filters_btn = WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        filters_btn.click()
                        print("✓ Filters panel opened")
                        time.sleep(4)
                        break
                    except:
                        continue
                
                date_input_selectors = [
                    "//input[@placeholder='mm/dd/yyyy hh:mm (a|p)m']",
                    "//input[@type='tel' and contains(@class, 'MuiInputBase-input')]"
                ]
                
                for selector in date_input_selectors:
                    try:
                        date_inputs = self.driver.find_elements(By.XPATH, selector)
                        for date_input in date_inputs:
                            if date_input.is_displayed():
                                date_input.click()
                                time.sleep(0.5)
                                date_input.send_keys('\ue009' + 'a')
                                time.sleep(0.3)
                                date_input.send_keys(yesterday_str)
                                print(f"✓ Date set: {yesterday_str}")
                                time.sleep(1)
                                break
                        break
                    except:
                        continue
                
                try:
                    apply_btn = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Apply')]"))
                    )
                    apply_btn.click()
                    print("✓ Apply clicked")
                    time.sleep(8)
                    self.wait_for_page_load(30)
                except:
                    pass
                print()
            except Exception as e:
                print(f"⚠ Date filter error: {e}\n")
            
            # Step 7
            self.show_status_window("⏳ Step 7/8: Exporting CSV...", "Exporting OrderBahn data to CSV...", 7)
            print("[7/7] Exporting CSV...")
            try:
                export_btn = WebDriverWait(self.driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'Export')]"))
                )
                export_btn.click()
                print("✓ Export clicked")
                time.sleep(2)
                
                csv_btn = WebDriverWait(self.driver, 8).until(
                    EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'CSV')]"))
                )
                csv_btn.click()
                print("✓ CSV selected")
                
                # Wait for download and track the file
                downloaded_file = self.wait_for_download(timeout=30)
                if downloaded_file:
                    self.orderbahn_csv = os.path.join(self.download_path, downloaded_file)
                    print(f"✓ OrderBahn CSV downloaded: {downloaded_file}")
                
                self.completed_steps.append("✅ OrderBahn CSV Export")
                self.show_status_window("✅ Step 7/8 Complete!", "OrderBahn CSV successfully downloaded! \n\nNext: ERP Automation (if applicable)", 7)
                
                return True, skip_mfa
                
            except Exception as e:
                print(f"❌ Export failed: {e}\n")
                self.failed_steps.append("❌ OrderBahn CSV Export")
                return False, skip_mfa
            
        except Exception as e:
            print(f"\n❌ OrderBahn automation failed: {e}\n")
            self.failed_steps.append("❌ OrderBahn Login/Navigation")
            return False, skip_mfa
    
    def run_erp_automation(self, tenant_name, erp_username, erp_password):
        """Run ERP automation with progress tracking"""
        try:
            print("\n" + "=" * 60)
            print(f"PART 2: {tenant_name.upper()} ERP AUTOMATION")
            print("=" * 60 + "\n")
            
            erp_url = self.erp_urls.get(tenant_name)
            if not erp_url:
                print(f"⚠ No ERP URL configured for tenant: {tenant_name}")
                print(f"⚠ Skipping ERP automation\n")
                self.completed_steps.append("⏭️ ERP Skipped (No URL configured)")
                return False
            
            self.show_status_window("⏳ Step 8/8: ERP Automation", "Starting ERP automation...\n\nLogging in to ERP system...", 8)
            
            self.close_driver()
            time.sleep(2)
            self.init_driver()
            
            print("[1/6] Logging in to ERP...")
            self.driver.get(erp_url)
            time.sleep(4)
            
            username_selectors = [
                "//input[@type='text']",
                "//input[@name='username']",
                "//input[@autocomplete='username']",
                "//input[@placeholder='Username']"
            ]
            
            username_field = None
            for selector in username_selectors:
                try:
                    username_field = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )
                    break
                except:
                    continue
            
            if not username_field:
                raise Exception("Could not find username field")
            
            username_field.clear()
            username_field.send_keys(erp_username)
            print("✓ Username entered")
            
            password_field = self.driver.find_element(By.XPATH, "//input[@type='password']")
            password_field.clear()
            password_field.send_keys(erp_password)
            print("✓ Password entered")
            
            login_btn = self.driver.find_element(By.XPATH, "//button[contains(., 'Login')]")
            login_btn.click()
            print("✓ Login submitted")
            
            time.sleep(8)
            self.wait_for_page_load(30)
            
            if "login" in self.driver.current_url.lower():
                raise Exception("Still on login page - ERP credentials may be incorrect")
            
            print("✓ Logged in successfully\n")
            
            print("[2/6] Navigating to Find AP Vouchers...")
            ap_url = erp_url.replace("/login", "/lo/accounting/ap/find")
            self.driver.get(ap_url)
            time.sleep(5)
            self.wait_for_page_load(30)
            print("✓ AP Voucher page loaded\n")
            
            print("[3/6] Selecting layout: QA_report...")
            try:
                time.sleep(3)
                
                layout_selectors = [
                    "//mat-select",
                    "//mat-select[contains(@class, 'mat-select')]",
                    "//select"
                ]
                
                layout_changed = False
                for selector in layout_selectors:
                    try:
                        layout_dropdown = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        layout_dropdown.click()
                        time.sleep(1)
                        
                        qa_option = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, "//mat-option//span[contains(text(),'QA_report')] | //option[text()='QA_report']"))
                        )
                        qa_option.click()
                        print("✓ Layout set to QA_report")
                        time.sleep(4)
                        self.wait_for_page_load(30)
                        layout_changed = True
                        break
                    except:
                        continue
                
                if not layout_changed:
                    print("⚠ Could not change layout - continuing with default\n")
                    
            except Exception as e:
                print(f"⚠ Layout change error: {e}\n")
            
            print("[4/6] Setting Create Date filter...")
            today = f"{datetime.now().month}/{datetime.now().day}/{datetime.now().year}"
            print(f"Using date: {today}")
            
            try:
                date_inputs = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "mat-date-range-input")
                    )
                )
                
                if len(date_inputs) >= 2:
                    start_input, end_input = date_inputs[0], date_inputs[1]
                    
                    start_input.click()
                    time.sleep(0.3)
                    start_input.send_keys(Keys.CONTROL + "a")
                    start_input.send_keys(today)
                    time.sleep(0.5)
                    
                    end_input.click()
                    time.sleep(0.3)
                    end_input.send_keys(Keys.CONTROL + "a")
                    end_input.send_keys(today)
                    time.sleep(0.5)
                    
                    print(f"✓ Date range set: {today} - {today}\n")
                else:
                    print("⚠ Could not find both date inputs\n")
                    
            except Exception as e:
                print(f"⚠ Could not set date: {e}\n")
            
            print("[5/6] Clicking 'Find Now'...")
            try:
                find_selectors = [
                    "//button[contains(@class,'mat-focus-indicator') and .//span[contains(text(),'Find Now')]]",
                    "//button[.//span[text()='Find Now']]",
                    "//button[contains(., 'Find Now')]",
                    "//span[text()='Find Now']/ancestor::button"
                ]
                
                find_clicked = False
                for selector in find_selectors:
                    try:
                        find_btn = WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, selector))
                        )
                        
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", find_btn)
                        time.sleep(0.5)
                        self.driver.execute_script("arguments[0].click();", find_btn)
                        print("✓ Find Now clicked")
                        find_clicked = True
                        break
                    except:
                        continue
                
                if not find_clicked:
                    print("⚠ Could not click Find Now - trying Enter key")
                    try:
                        date_inputs = self.driver.find_elements(By.CSS_SELECTOR, "mat-date-range-input input")
                        if date_inputs:
                            date_inputs[-1].send_keys(Keys.ENTER)
                            print("✓ Triggered search via Enter key")
                    except:
                        print("❌ Could not trigger Find Now")
                
                time.sleep(8)
                self.wait_for_page_load(30)
                print()
                
            except Exception as e:
                print(f"⚠ Find Now error: {e}\n")
            
            print("[6/6] Exporting results as CSV...")
            try:
                time.sleep(2)
                
                csv_clicked = False
                
                csv_selectors = [
                    "//fa-icon[@mattooltip='Export CSV']/parent::button",
                    "//button[@mattooltip='Export CSV']",
                    "//fa-icon[contains(@class, 'ng-fa-icon')][@mattooltip='Export CSV']/ancestor::button",
                    "//button[.//fa-icon[@mattooltip='Export CSV']]",
                ]
                
                for selector in csv_selectors:
                    try:
                        csv_btn = WebDriverWait(self.driver, 3).until(
                            EC.presence_of_element_located((By.XPATH, selector))
                        )
                        
                        if csv_btn.is_displayed():
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", csv_btn)
                            time.sleep(0.5)
                            self.driver.execute_script("arguments[0].click();", csv_btn)
                            csv_clicked = True
                            break
                    except:
                        continue
                
                if not csv_clicked:
                    try:
                        fa_icons = self.driver.find_elements(By.TAG_NAME, "fa-icon")
                        for icon in fa_icons:
                            try:
                                tooltip = icon.get_attribute("mattooltip") or ""
                                if "csv" in tooltip.lower():
                                    parent_btn = icon.find_element(By.XPATH, "./ancestor::button")
                                    if parent_btn.is_displayed():
                                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", parent_btn)
                                        time.sleep(0.5)
                                        self.driver.execute_script("arguments[0].click();", parent_btn)
                                        csv_clicked = True
                                        break
                            except:
                                continue
                    except:
                        pass
                
                if csv_clicked:
                    print(f"✓ CSV export button clicked")
                    
                    downloaded_file = self.wait_for_download(timeout=30)
                    
                    if downloaded_file:
                        self.erp_csv = os.path.join(self.download_path, downloaded_file)
                        file_size = os.path.getsize(self.erp_csv)
                        
                        print(f"✓ CSV downloaded successfully!")
                        print(f"  Filename: {downloaded_file}")
                        print(f"  Size: {file_size:,} bytes")
                        print(f"  Location: {self.erp_csv}\n")
                        
                        self.completed_steps.append("✅ ERP CSV Export")
                    else:
                        print("⚠ Download did not complete within 30 seconds")
                        print(f"  Check folder manually: {self.download_path}\n")
                        self.failed_steps.append("⚠ ERP CSV Export (Timeout)")
                else:
                    print("❌ Could not find CSV export button\n")
                    self.failed_steps.append("❌ ERP CSV Export (Button not found)")
                    
            except Exception as e:
                print(f"❌ CSV export error: {e}\n")
                self.failed_steps.append(f"❌ ERP CSV Export: {str(e)[:30]}")
            
            self.show_status_window("✅ Step 8/8 Complete!", "ERP automation finished!\n\nGenerating Excel report...", 8)
            
            print("=" * 60)
            print(f"✅ {tenant_name.upper()} ERP AUTOMATION COMPLETE!")
            print("=" * 60 + "\n")
            
            return True
            
        except Exception as e:
            print(f"\n❌ ERP automation failed: {e}\n")
            self.failed_steps.append("❌ ERP Login/Navigation")
            return False
    
    def transfer_data_to_excel(self, tenant_name):
        """Transfer data from CSVs to Excel file with formulas and conditional formatting"""
        try:
            print("\n" + "=" * 60)
            print("PART 3: DATA TRANSFER TO EXCEL")
            print("=" * 60 + "\n")
            
            # Generate filename
            date_str = datetime.now().strftime("%m%d%Y")
            excel_filename = f"{tenant_name} QA Report AP {date_str}.xlsx"
            excel_path = os.path.join(self.download_path, excel_filename)
            
            print(f"[1/4] Creating Excel file: {excel_filename}")
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            sheets = {
                'Core_Data': wb.create_sheet("Core_Data"),
                'OB_Data': wb.create_sheet("OB_Data"),
                'Core_Original': wb.create_sheet("Core_Original"),
                'OB_Original': wb.create_sheet("OB_Original")
            }
            
            print("✓ Created 4 sheets\n")
            
            # Define styles
            self._define_styles(wb)
            
            # [2/4] Process Core data
            if self.erp_csv and os.path.exists(self.erp_csv):
                print(f"[2/4] Processing Core data...")
                df_core = pd.read_csv(self.erp_csv)
                
                # Write Core_Original
                self._write_original_data(sheets['Core_Original'], df_core, 'Core')
                print("✓ Core_Original written")
                
                # Write Core_Data with formulas
                self._write_core_data_with_formulas(sheets['Core_Data'], df_core)
                print("✓ Core_Data written with formulas and conditional formatting\n")
                self.completed_steps.append("✅ Core Data Processing")
            else:
                print("[2/4] No ERP CSV found\n")
                self.completed_steps.append("⏭️ Core Data Processing (No CSV)")
            
            # [3/4] Process OB data
            if self.orderbahn_csv and os.path.exists(self.orderbahn_csv):
                print(f"[3/4] Processing OB data...")
                df_ob = pd.read_csv(self.orderbahn_csv)
                
                # Write OB_Original
                self._write_original_data(sheets['OB_Original'], df_ob, 'OB')
                print("✓ OB_Original written")
                
                # Write OB_Data with formulas and summary
                self._write_ob_data_with_formulas(sheets['OB_Data'], df_ob)
                print("✓ OB_Data written with formulas and conditional formatting\n")
                self.completed_steps.append("✅ OB Data Processing")
            else:
                print("[3/4] No OB CSV found\n")
                self.completed_steps.append("⏭️ OB Data Processing (No CSV)")
            
            # [4/4] Save workbook
            print(f"[4/4] Saving Excel file...")
            wb.save(excel_path)
            
            file_size = os.path.getsize(excel_path)
            print(f"✓ Excel file saved successfully!")
            print(f"  Filename: {excel_filename}")
            print(f"  Size: {file_size:,} bytes")
            print(f"  Location: {excel_path}\n")
            
            self.completed_steps.append("✅ Excel Report Generation")
            
            print("=" * 60)
            print("✅ DATA TRANSFER COMPLETE!")
            print("=" * 60 + "\n")
            
            return excel_path
            
        except Exception as e:
            print(f"\n❌ Data transfer failed: {e}\n")
            import traceback
            traceback.print_exc()
            self.failed_steps.append(f"❌ Excel Generation: {str(e)[:30]}")
            return None

    def _define_styles(self, wb):
        """Define named styles for headers and cells"""
        from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
        
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(color="FFFFFF", bold=True, size=11)
        header_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        wb.add_named_style(header_style)

    def _write_original_data(self, ws, df, data_type):
        """Write original data with double headers"""
        if data_type == 'Core':
            headers = [
                'Voucher #', 'Applied Date', 'Vendor', 'Vendor Name', 'Remit Vendor',
                'Invoice #', 'PO #', 'Posted By', 'Invoice Date', 'Payment Terms',
                'Invoice Amount', 'Discount', 'Net Amount', 'Order #', 'Status',
                'Payment Date', 'Project #', 'Customer Acct', 'Company', 'Pay Next',
                'Due Date', 'Check #', 'Type', 'Payment Type', 'Hold', 'Void Date', 'Attachments'
            ]
        else:  # OB
            headers = [
                'VendorName', 'PONumber', 'InvoiceNumber', 'InvoiceDate', 'Due Date',
                'InvoiceAmount', 'SalesTax', 'Freight', 'Surcharge', 'DiscountTerms',
                'Terms', 'ArchiveDate', 'PostingDate', 'Record Status', 'Comments'
            ]
        
        # Write double headers (rows 1-2)
        for row in [1, 2]:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.style = 'header_style'
        
        # Write data starting at row 3
        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Adjust column widths
        self._adjust_column_widths(ws)

    def _write_core_data_with_formulas(self, ws, df):
        """Write Core_Data with formulas and conditional formatting"""
        headers = [
            'Clean Invoice #', 'Vendor Name', 'Invoice #', 'OB Invoice Lookup', 'PO #',
            'Order #', 'Invoice Date', 'Due Date', 'Payment Terms', 'Invoice Amount',
            'Net Amount', 'OB Amount', 'Discount Amount', 'Discount%', 'Amount Diff',
            'Ob Comment', 'Posting Date', 'Payment date', 'Days', 'User', 'Attachments'
        ]
        
        # Write double headers (rows 1-2)
        for row in [1, 2]:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.style = 'header_style'
        
        # Write data with formulas starting at row 3
        data_start_row = 3
        for r_idx, row in enumerate(df.itertuples(index=False), start=data_start_row):
            # Static values and formulas
            ws.cell(row=r_idx, column=1, value=f"=Core_Original!F{r_idx}")  # Clean Invoice #
            ws.cell(row=r_idx, column=2, value=f"=Core_Original!D{r_idx}")  # Vendor Name
            ws.cell(row=r_idx, column=3, value=f"=Core_Original!F{r_idx}")  # Invoice #
            ws.cell(row=r_idx, column=4, value=f"=Core_Original!F{r_idx}")  # OB Invoice Lookup
            ws.cell(row=r_idx, column=5, value=f"=Core_Original!G{r_idx}")  # PO #
            ws.cell(row=r_idx, column=6, value=f"=Core_Original!N{r_idx}")  # Order #
            ws.cell(row=r_idx, column=7, value=f"=Core_Original!I{r_idx}")  # Invoice Date
            ws.cell(row=r_idx, column=8, value=f"=Core_Original!U{r_idx}")  # Due Date
            ws.cell(row=r_idx, column=9, value=f"=Core_Original!J{r_idx}")  # Payment Terms
            ws.cell(row=r_idx, column=10, value=f"=Core_Original!K{r_idx}")  # Invoice Amount
            ws.cell(row=r_idx, column=11, value=f"=Core_Original!M{r_idx}")  # Net Amount
            ws.cell(row=r_idx, column=12, value=f"=Core_Original!M{r_idx}")  # OB Amount
            ws.cell(row=r_idx, column=13, value=f"=Core_Original!L{r_idx}")  # Discount Amount
            ws.cell(row=r_idx, column=14, value=f"=IF(K{r_idx}=0,0,M{r_idx}/K{r_idx})")  # Discount%
            ws.cell(row=r_idx, column=15, value=0)  # Amount Diff (static 0)
            ws.cell(row=r_idx, column=16, value="")  # Ob Comment
            ws.cell(row=r_idx, column=17, value=f"=Core_Original!B{r_idx}")  # Posting Date
            ws.cell(row=r_idx, column=18, value=f"=Core_Original!O{r_idx}")  # Payment date
            ws.cell(row=r_idx, column=19, value=f"=IF(AND(R{r_idx}<>'',Q{r_idx}<>''),R{r_idx}-Q{r_idx},'')")  # Days
            ws.cell(row=r_idx, column=20, value=f"=Core_Original!H{r_idx}")  # User
            ws.cell(row=r_idx, column=21, value=f"=IF(Core_Original!AA{r_idx}<>'','Yes','')")  # Attachments
        
        # Add conditional formatting for Amount Diff column (O)
        self._add_conditional_formatting(ws, 'O', data_start_row, len(df) + 2)
        
        # Adjust column widths
        self._adjust_column_widths(ws)

    def _write_ob_data_with_formulas(self, ws, df):
        """Write OB_Data with formulas, summary row, and conditional formatting"""
        headers = [
            'Clean Invoice #', 'Vendor Name', 'Lookup vendor name', 'PO Number', 'Invoice Number',
            'Lookup invoice Number', 'Invoice Date', 'Lookup invoice date', 'Due Date',
            'Lookup Due date', 'Invoice Amount', 'Lookup invoice Total', 'Sales Tax',
            'Freight', 'Surcharge', 'Discount Terms', 'Terms', 'Lookup terms', 'Days',
            'Check date flag', 'check amount flag', 'Check Due Date', 'Check vendor',
            'Check terms', 'Comment'
        ]
        
        # Summary row (row 1)
        total_formula = f"=SUM(K3:K{len(df)+2})"
        ws.cell(row=1, column=1, value="")  # Empty
        ws.cell(row=1, column=11, value=total_formula)  # Total Invoice Amount
        ws.cell(row=1, column=12, value=total_formula)  # Total Lookup Amount
        
        # Write double headers (rows 2-3)
        for row in [2, 3]:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.style = 'header_style'
        
        # Write data with formulas starting at row 4
        data_start_row = 4
        for r_idx, row in enumerate(df.itertuples(index=False), start=data_start_row):
            base_idx = r_idx - 1  # Adjust for OB_Original row reference
            
            # Static values and formulas
            ws.cell(row=r_idx, column=1, value=f"=OB_Original!C{base_idx}")  # Clean Invoice #
            ws.cell(row=r_idx, column=2, value=f"=OB_Original!A{base_idx}")  # Vendor Name
            ws.cell(row=r_idx, column=3, value=f"=OB_Original!A{base_idx}")  # Lookup vendor name
            ws.cell(row=r_idx, column=4, value=f"=OB_Original!B{base_idx}")  # PO Number
            ws.cell(row=r_idx, column=5, value=f"=OB_Original!C{base_idx}")  # Invoice Number
            ws.cell(row=r_idx, column=6, value=f"=OB_Original!C{base_idx}")  # Lookup invoice Number
            ws.cell(row=r_idx, column=7, value=f"=OB_Original!D{base_idx}")  # Invoice Date
            ws.cell(row=r_idx, column=8, value=f"=OB_Original!D{base_idx}")  # Lookup invoice date
            ws.cell(row=r_idx, column=9, value=f"=OB_Original!E{base_idx}")  # Due Date
            ws.cell(row=r_idx, column=10, value=f"=OB_Original!E{base_idx}")  # Lookup Due date
            ws.cell(row=r_idx, column=11, value=f"=OB_Original!F{base_idx}")  # Invoice Amount
            ws.cell(row=r_idx, column=12, value=f"=OB_Original!F{base_idx}")  # Lookup invoice Total
            ws.cell(row=r_idx, column=13, value=f"=OB_Original!G{base_idx}")  # Sales Tax
            ws.cell(row=r_idx, column=14, value=f"=OB_Original!H{base_idx}")  # Freight
            ws.cell(row=r_idx, column=15, value=f"=OB_Original!I{base_idx}")  # Surcharge
            ws.cell(row=r_idx, column=16, value=f"=OB_Original!J{base_idx}")  # Discount Terms
            ws.cell(row=r_idx, column=17, value=f"=OB_Original!K{base_idx}")  # Terms
            ws.cell(row=r_idx, column=18, value=f"=OB_Original!K{base_idx}")  # Lookup terms
            ws.cell(row=r_idx, column=19, value="")  # Days - extract from terms
            ws.cell(row=r_idx, column=20, value=f"=IF(G{r_idx}<>'',1,0)")  # Check date flag
            ws.cell(row=r_idx, column=21, value=0)  # check amount flag (default 0)
            ws.cell(row=r_idx, column=22, value=0)  # Check Due Date (default 0)
            ws.cell(row=r_idx, column=23, value=1)  # Check vendor (default 1)
            ws.cell(row=r_idx, column=24, value=1)  # Check terms (default 1)
            ws.cell(row=r_idx, column=25, value=f"=OB_Original!O{base_idx}")  # Comment
        
        # Add conditional formatting for flag columns (T, U, V, W, X)
        for col in ['T', 'U', 'V', 'W', 'X']:
            self._add_flag_conditional_formatting(ws, col, data_start_row, len(df) + 3)
        
        # Adjust column widths
        self._adjust_column_widths(ws)

    def _add_conditional_formatting(self, ws, column, start_row, end_row):
        """Add conditional formatting for amount differences (non-zero = red)"""
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.conditional_formatting.add(
            f'{column}{start_row}:{column}{end_row}',
            CellIsRule(operator='notEqual', formula=['0'], fill=red_fill)
        )

    def _add_flag_conditional_formatting(self, ws, column, start_row, end_row):
        """Add conditional formatting for flag columns (0=red, 1=green)"""
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Green for 1
        ws.conditional_formatting.add(
            f'{column}{start_row}:{column}{end_row}',
            CellIsRule(operator='equal', formula=['1'], fill=green_fill)
        )
        
        # Red for 0
        ws.conditional_formatting.add(
            f'{column}{start_row}:{column}{end_row}',
            CellIsRule(operator='equal', formula=['0'], fill=red_fill)
        )

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def run(self):
        """Main execution flow with progress popups"""
        try:
            print("=" * 60)
            print("UNIFIED ORDERBAHN & ERP AUTOMATION")
            print("=" * 60 + "\n")
            
            cache = self.load_cache()
            
            setup = self.show_initial_setup_dialog(cache)
            
            if setup['cancelled']:
                print("❌ Automation cancelled by user\n")
                return
            
            orderbahn_email = setup['orderbahn_email']
            orderbahn_password = setup['orderbahn_password']
            tenant_name = setup['tenant']
            erp_username = setup['erp_username']
            erp_password = setup['erp_password']
            save_credentials = setup['save_credentials']
            browser = setup['browser']
            
            print(f"📋 Configuration:")
            print(f"   Email: {orderbahn_email}")
            print(f"   Tenant: {tenant_name}")
            print(f"   Browser: {browser.title()}")
            if erp_username:
                print(f"   ERP User: {erp_username}")
            print()
            
            self.init_driver(browser)
            
            # Run OrderBahn automation
            orderbahn_success, skip_mfa = self.run_orderbahn_automation(
                orderbahn_email, 
                orderbahn_password, 
                tenant_name,
                cache.get('skip_mfa', True)
            )
            
            if not orderbahn_success:
                print("❌ OrderBahn automation failed")
                self.failed_steps.append("❌ OrderBahn Automation (Critical)")
                self.show_final_summary()
                return
            
            # Run ERP automation if applicable
            should_run_erp = False
            if tenant_name in self.erp_urls:
                if erp_username and erp_password:
                    should_run_erp = True
                    print(f"\n🔔 Running ERP automation for {tenant_name}...")
                else:
                    print(f"\n⏭ Skipping ERP automation (no credentials provided)\n")
                    self.completed_steps.append("⏭️ ERP Automation (No Credentials)")
            else:
                print(f"\n⏭ No ERP system configured for {tenant_name}\n")
                self.completed_steps.append("⏭️ ERP Automation (No ERP URL)")
            
            erp_success = False
            if should_run_erp:
                erp_success = self.run_erp_automation(tenant_name, erp_username, erp_password)
            
            # Close browser before Excel processing
            self.close_driver()
            
            # Transfer data to Excel
            excel_path = self.transfer_data_to_excel(tenant_name)
            
            # Save credentials if requested
            if save_credentials:
                self.save_cache(
                    orderbahn_email, 
                    orderbahn_password,
                    erp_username if should_run_erp else None,
                    erp_password if should_run_erp else None,
                    tenant_name,
                    skip_mfa
                )
            
            # Show final summary with all steps
            self.show_final_summary()
            
        except Exception as e:
            print(f"\n❌ Automation failed: {e}\n")
            import traceback
            traceback.print_exc()
            self.failed_steps.append(f"❌ Unexpected Error: {str(e)[:40]}")
            self.show_final_summary()
        
        finally:
            if self.driver:
                input("\nPress Enter to close...")
                self.close_driver()


if __name__ == "__main__":
    automation = UnifiedAutomation()
    automation.run()
